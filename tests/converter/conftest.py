#!/usr/bin/env python3
# tests/converter/conftest.py
"""Shared fixtures for all TrajCenter test modules.

Author: Clement RACINET

Provides low-level fixtures (DataFrames, metadata) reusable by
``test_trajectory.py``, ``test_mod_converter.py``,
``test_excel_converter.py``, ``test_csv_converter.py``, etc.
"""

from __future__ import annotations

from pathlib import Path
from textwrap import dedent

import pytest
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Fixtures — synthetic .mod files
# ---------------------------------------------------------------------------


@pytest.fixture
def mod_simple(tmp_path: Path) -> Path:
    """Minimal .mod file with two MoveL instructions and a variable speed."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
                MoveL [[150.0,250.0,350.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "simple.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_with_literal_speed(tmp_path: Path) -> Path:
    """Minimal .mod file with a literal RAPID speed value (v500)."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                MoveL [[10.0,20.0,30.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],v500,z10,Tool_formage\\wobj:=Wobj_SerreFlan;
                MoveJ [[40.0,50.0,60.0],[1.0,0.0,0.0,0.0],[-1,0,1,0],[9E9,9E9,9E9,9E9,9E9,9E9]],v1000,fine,Tool_formage\\wobj:=Wobj_SerreFlan;
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "literal_speed.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_with_eax(tmp_path: Path) -> Path:
    """Minimal .mod file with one active external axis (eax_a = 45.0)."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[45.0,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "eax.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_multiline(tmp_path: Path) -> Path:
    """Minimal .mod file with a robtarget formatted across multiple lines."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                MoveL [[100.0,200.0,300.0],
                       [1.0,0.0,0.0,0.0],
                       [0,0,0,0],
                       [9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "multiline.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_multiple_tools(tmp_path: Path) -> Path:
    """Minimal .mod file with two distinct tools and two distinct wobjs."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                MoveL [[1.0,2.0,3.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_A\\wobj:=Wobj_A;
                MoveL [[4.0,5.0,6.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_B\\wobj:=Wobj_B;
                MoveL [[7.0,8.0,9.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_A\\wobj:=Wobj_A;
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "multi_tools.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_empty(tmp_path: Path) -> Path:
    """Minimal .mod file with no Move instructions."""
    content = dedent("""\
        MODULE TestModule
            PROC TestProc()
                ! No Move instruction here
            ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "empty.mod"
    p.write_text(content, encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# Fixtures — synthetic Excel workbooks
# ---------------------------------------------------------------------------


def _make_xlsx(path: Path, sheets: dict[str, list[dict]]) -> Path:
    """Create an ``.xlsx`` file from a ``{sheet_name: [rows]}`` mapping.

    The first row of each sheet is used as the header (keys of the first
    dict in the list).

    Args:
        path: Destination file path.
        sheets: Mapping of sheet name to a list of row dicts.

    Returns:
        The path to the created ``.xlsx`` file.
    """
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        assert ws is not None
        ws.title = sheet_name
        first = False
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    wb.save(path)
    return path


@pytest.fixture
def xlsx_simple(tmp_path: Path) -> Path:
    """Minimal workbook: one sheet with XYZ and quaternion columns."""
    return _make_xlsx(
        tmp_path / "simple.xlsx",
        {
            "traj": [
                {
                    "x": 100.0,
                    "y": 200.0,
                    "z": 300.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                },
                {
                    "x": 150.0,
                    "y": 250.0,
                    "z": 350.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                },
            ],
        },
    )


@pytest.fixture
def xlsx_xyz_only(tmp_path: Path) -> Path:
    """Workbook with XYZ columns only — identity orientation applied by default."""
    return _make_xlsx(
        tmp_path / "xyz_only.xlsx",
        {
            "traj": [
                {"x": 10.0, "y": 20.0, "z": 30.0},
                {"x": 40.0, "y": 50.0, "z": 60.0},
            ],
        },
    )


@pytest.fixture
def xlsx_aliases(tmp_path: Path) -> Path:
    """Workbook with non-canonical column names (aliases, accents, mixed case)."""
    return _make_xlsx(
        tmp_path / "aliases.xlsx",
        {
            "traj": [
                {
                    "PosX": 1.0,
                    "PosY": 2.0,
                    "PosZ": 3.0,
                    "Vitesse": "v500",
                    "Répère": "Wobj_A",
                    "Outil": "Tool_A",
                },
            ],
        },
    )


@pytest.fixture
def xlsx_multi_traj(tmp_path: Path) -> Path:
    """Workbook with two trajectory sheets."""
    return _make_xlsx(
        tmp_path / "multi_traj.xlsx",
        {
            "traj_A": [
                {"x": 1.0, "y": 2.0, "z": 3.0},
            ],
            "traj_B": [
                {"x": 4.0, "y": 5.0, "z": 6.0},
                {"x": 7.0, "y": 8.0, "z": 9.0},
            ],
        },
    )


@pytest.fixture
def xlsx_with_tools_sheet(tmp_path: Path) -> Path:
    """Workbook with a trajectory sheet, a tools sheet and a wobjs sheet."""
    return _make_xlsx(
        tmp_path / "with_refs.xlsx",
        {
            "traj": [
                {"x": 1.0, "y": 2.0, "z": 3.0, "tool": "Tool_A", "wobj": "Wobj_A"},
                {"x": 4.0, "y": 5.0, "z": 6.0, "tool": "Tool_B", "wobj": "Wobj_A"},
            ],
            "tools": [
                {"name": "Tool_A"},
                {"name": "Tool_B"},
            ],
            "wobjs": [
                {"name": "Wobj_A"},
            ],
        },
    )


@pytest.fixture
def xlsx_missing_xyz(tmp_path: Path) -> Path:
    """Workbook without XYZ columns — must raise ``ValueError``."""
    return _make_xlsx(
        tmp_path / "missing_xyz.xlsx",
        {
            "traj": [
                {"speed": "v500", "zone": "z0"},
            ],
        },
    )


@pytest.fixture
def xlsx_with_meta_sheet(tmp_path: Path) -> Path:
    """Workbook with a meta sheet — must be silently ignored."""
    return _make_xlsx(
        tmp_path / "with_meta.xlsx",
        {
            "traj": [
                {"x": 1.0, "y": 2.0, "z": 3.0},
            ],
            "meta": [
                {"key": "author", "value": "test"},
            ],
        },
    )


@pytest.fixture
def xlsx_empty_rows(tmp_path: Path) -> Path:
    """Workbook with fully empty rows interspersed between data rows."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "traj"
    ws.append(["x", "y", "z"])
    ws.append([1.0, 2.0, 3.0])
    ws.append([None, None, None])  # empty row
    ws.append([4.0, 5.0, 6.0])
    wb.save(tmp_path / "empty_rows.xlsx")
    return tmp_path / "empty_rows.xlsx"


# ---------------------------------------------------------------------------
# Helpers — synthetic CSV files
# ---------------------------------------------------------------------------


def _write_csv(path: Path, content: str, encoding: str = "utf-8") -> Path:
    """Write a synthetic CSV file and return its path.

    Args:
        path: Destination file path.
        content: Raw CSV content as a string.
        encoding: File encoding (default: ``"utf-8"``).

    Returns:
        The path to the written CSV file.
    """
    path.write_text(content, encoding=encoding)
    return path


# ---------------------------------------------------------------------------
# Fixtures — synthetic CSV files
# ---------------------------------------------------------------------------


@pytest.fixture
def csv_simple(tmp_path: Path) -> Path:
    """Minimal CSV: XYZ + quaternions, comma separator."""
    return _write_csv(
        tmp_path / "simple.csv",
        "x,y,z,q1,q2,q3,q4\n"
        "100.0,200.0,300.0,1.0,0.0,0.0,0.0\n"
        "150.0,250.0,350.0,1.0,0.0,0.0,0.0\n",
    )


@pytest.fixture
def csv_semicolon(tmp_path: Path) -> Path:
    """CSV with semicolon separator (French Excel export)."""
    return _write_csv(
        tmp_path / "semicolon.csv",
        "x;y;z;q1;q2;q3;q4\n"
        "10.0;20.0;30.0;1.0;0.0;0.0;0.0\n"
        "40.0;50.0;60.0;1.0;0.0;0.0;0.0\n",
    )


@pytest.fixture
def csv_xyz_only(tmp_path: Path) -> Path:
    """CSV with XYZ columns only — identity orientation applied by default."""
    return _write_csv(
        tmp_path / "xyz_only.csv",
        "x,y,z\n10.0,20.0,30.0\n40.0,50.0,60.0\n",
    )


@pytest.fixture
def csv_aliases(tmp_path: Path) -> Path:
    """CSV with non-canonical column names (aliases, mixed case)."""
    return _write_csv(
        tmp_path / "aliases.csv",
        "PosX,PosY,PosZ,VITESSE\n1.0,2.0,3.0,v500\n",
    )


@pytest.fixture
def csv_with_tools(tmp_path: Path) -> Path:
    """CSV with tool and wobj columns."""
    return _write_csv(
        tmp_path / "with_tools.csv",
        "x,y,z,tool,wobj\n1.0,2.0,3.0,Tool_A,Wobj_A\n4.0,5.0,6.0,Tool_B,Wobj_A\n",
    )


@pytest.fixture
def csv_missing_xyz(tmp_path: Path) -> Path:
    """CSV without XYZ columns — must raise ``ValueError``."""
    return _write_csv(
        tmp_path / "missing_xyz.csv",
        "speed,zone\nv500,z0\n",
    )


@pytest.fixture
def csv_empty_rows(tmp_path: Path) -> Path:
    """CSV with fully empty rows interspersed between data rows."""
    return _write_csv(
        tmp_path / "empty_rows.csv",
        "x,y,z\n1.0,2.0,3.0\n,,\n4.0,5.0,6.0\n",
    )


@pytest.fixture
def csv_with_bom(tmp_path: Path) -> Path:
    """CSV encoded as UTF-8 with BOM (Windows Excel export)."""
    return _write_csv(
        tmp_path / "bom.csv",
        "x,y,z\n1.0,2.0,3.0\n",
        encoding="utf-8-sig",
    )


@pytest.fixture
def csv_full(tmp_path: Path) -> Path:
    """CSV with all canonical columns present."""
    return _write_csv(
        tmp_path / "full.csv",
        "x,y,z,q1,q2,q3,q4,move_type,speed,zone,tool,wobj\n"
        "100.0,200.0,300.0,1.0,0.0,0.0,0.0,MoveL,v500,z10,Tool_formage,Wobj_SerreFlan\n"
        "150.0,250.0,350.0,1.0,0.0,0.0,0.0,MoveJ,v1000,fine,Tool_formage,Wobj_SerreFlan\n",
    )


@pytest.fixture
def xlsx_with_full_meta(tmp_path: Path) -> Path:
    """Workbook with a complete meta sheet (name, robot_model, custom field)."""
    return _make_xlsx(
        tmp_path / "full_meta.xlsx",
        {
            "traj": [{"x": 1.0, "y": 2.0, "z": 3.0}],
            "meta": [
                {"key": "name", "value": "Trajectoire_Soudure"},
                {"key": "robot_model", "value": "IRB6700-205/2.80"},
                {"key": "author", "value": "Jean Dupont"},
            ],
        },
    )


@pytest.fixture
def mod_bad_confdata(tmp_path: Path) -> Path:
    """A .mod file with malformed confdata (non-integer values)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[X,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "bad_confdata.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_zone_var(tmp_path: Path) -> Path:
    """A .mod file with a zone specified as a variable name (not a RAPID literal)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,ma_zone,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "zone_var.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_no_robtarget(tmp_path: Path) -> Path:
    """A .mod file where a Move instruction has no inline robtarget."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL pTarget1,vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "no_robtarget.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def csv_unknown_col(tmp_path: Path) -> Path:
    """CSV with one unknown column alongside valid XYZ columns."""
    return _write_csv(
        tmp_path / "unknown_col.csv",
        "x,y,z,custom_col\n1.0,2.0,3.0,99\n",
    )
