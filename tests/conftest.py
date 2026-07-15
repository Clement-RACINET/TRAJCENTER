#!/usr/bin/env python3
# tests/converter/conftest.py
"""Shared fixtures for all TrajCenter test modules.

Author: Clement RACINET

Provides low-level fixtures (DataFrames, metadata) reusable by
``test_trajectory.py``, ``test_mod_converter.py``,
``test_excel_converter.py``, ``test_csv_converter.py``, etc.
"""

from __future__ import annotations

from pathlib import Path
from textwrap import dedent

import pandas as pd
import pytest
from openpyxl import Workbook

from trajcenter.core.trajectory import Trajectory, TrajectoryMeta


# ---------------------------------------------------------------------------
# Fixtures — synthetic .mod files
# ---------------------------------------------------------------------------


@pytest.fixture
def mod_simple(tmp_path: Path) -> Path:
    """Minimal .mod file with two MoveL instructions and a variable speed."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
          MoveL [[150.0,250.0,350.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "simple.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_with_literal_speed(tmp_path: Path) -> Path:
    """Minimal .mod file with a literal RAPID speed value (v500)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[10.0,20.0,30.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],v500,z10,Tool_formage\\wobj:=Wobj_SerreFlan;
          MoveJ [[40.0,50.0,60.0],[1.0,0.0,0.0,0.0],[-1,0,1,0],[9E9,9E9,9E9,9E9,9E9,9E9]],v1000,fine,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "literal_speed.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_with_eax(tmp_path: Path) -> Path:
    """Minimal .mod file with one active external axis (eax_a = 45.0)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[45.0,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "eax.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_multiline(tmp_path: Path) -> Path:
    """Minimal .mod file with a robtarget formatted across multiple lines."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],
                 [1.0,0.0,0.0,0.0],
                 [0,0,0,0],
                 [9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "multiline.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_multiple_tools(tmp_path: Path) -> Path:
    """Minimal .mod file with two distinct tools and two distinct wobjs."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[1.0,2.0,3.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_A\\wobj:=Wobj_A;
          MoveL [[4.0,5.0,6.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_B\\wobj:=Wobj_B;
          MoveL [[7.0,8.0,9.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_A\\wobj:=Wobj_A;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "multi_tools.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_empty(tmp_path: Path) -> Path:
    """Minimal .mod file with no Move instructions."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          ! No Move instruction here
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "empty.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_bad_confdata(tmp_path: Path) -> Path:
    """A .mod file with malformed confdata (non-integer values)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[X,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "bad_confdata.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_zone_var(tmp_path: Path) -> Path:
    """A .mod file with a zone specified as a variable name (not a RAPID literal)."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL [[100.0,200.0,300.0],[1.0,0.0,0.0,0.0],[0,0,0,0],[9E9,9E9,9E9,9E9,9E9,9E9]],vitesse,ma_zone,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "zone_var.mod"
    p.write_text(content, encoding="utf-8")
    return p


@pytest.fixture
def mod_no_robtarget(tmp_path: Path) -> Path:
    """A .mod file where a Move instruction has no inline robtarget."""
    content = dedent("""\
        MODULE TestModule
        PROC TestProc()
          MoveL pTarget1,vitesse,z0,Tool_formage\\wobj:=Wobj_SerreFlan;
        ENDPROC
        ENDMODULE
    """)
    p = tmp_path / "no_robtarget.mod"
    p.write_text(content, encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# Fixtures — synthetic Excel workbooks
# ---------------------------------------------------------------------------


def _make_xlsx(path: Path, sheets: dict[str, list[dict]]) -> Path:
    """Create an ``.xlsx`` file from a dict of sheet_name → list of row dicts.

    Args:
        path: Destination path for the workbook.
        sheets: Mapping of sheet name to list of row dicts.

    Returns:
        The path to the created workbook.
    """
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name  # type: ignore[union-attr]
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        if rows:
            ws.append(list(rows[0].keys()))  # type: ignore[union-attr]
            for row in rows:
                ws.append(list(row.values()))  # type: ignore[union-attr]
    wb.save(path)
    return path


@pytest.fixture
def xlsx_simple(tmp_path: Path) -> Path:
    """Simple ``.xlsx`` with 2 complete trajectory rows on a default sheet."""
    return _make_xlsx(
        tmp_path / "simple.xlsx",
        {
            "Sheet1": [
                {
                    "x": 100.0,
                    "y": 200.0,
                    "z": 300.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                    "move_type": "MoveL",
                    "speed": "v500",
                    "zone": "z10",
                    "tool": "Tool_formage",
                    "wobj": "Wobj_SerreFlan",
                },
                {
                    "x": 150.0,
                    "y": 250.0,
                    "z": 350.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                    "move_type": "MoveL",
                    "speed": "v500",
                    "zone": "z10",
                    "tool": "Tool_formage",
                    "wobj": "Wobj_SerreFlan",
                },
            ]
        },
    )


@pytest.fixture
def xlsx_missing_xyz(tmp_path: Path) -> Path:
    """An ``.xlsx`` with only quaternion columns — XYZ missing."""
    return _make_xlsx(
        tmp_path / "missing_xyz.xlsx",
        {"Sheet1": [{"q1": 1.0, "q2": 0.0, "q3": 0.0, "q4": 0.0}]},
    )


@pytest.fixture
def xlsx_xyz_only(tmp_path: Path) -> Path:
    """An ``.xlsx`` with only XYZ columns — quaternions absent."""
    return _make_xlsx(
        tmp_path / "xyz_only.xlsx",
        {"Sheet1": [{"x": 1.0, "y": 2.0, "z": 3.0}]},
    )


@pytest.fixture
def xlsx_aliases(tmp_path: Path) -> Path:
    """An ``.xlsx`` using column aliases (PosX, vitesse) instead of canonical names."""
    return _make_xlsx(
        tmp_path / "aliases.xlsx",
        {
            "Sheet1": [
                {
                    "PosX": 1.0,
                    "PosY": 2.0,
                    "PosZ": 3.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                    "vitesse": "v500",
                }
            ]
        },
    )


@pytest.fixture
def xlsx_multi_traj(tmp_path: Path) -> Path:
    """An ``.xlsx`` with two trajectory sheets (traj_A and traj_B)."""
    return _make_xlsx(
        tmp_path / "multi_traj.xlsx",
        {
            "traj_A": [
                {
                    "x": 1.0,
                    "y": 2.0,
                    "z": 3.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                }
            ],
            "traj_B": [
                {
                    "x": 4.0,
                    "y": 5.0,
                    "z": 6.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                },
                {
                    "x": 7.0,
                    "y": 8.0,
                    "z": 9.0,
                    "q1": 1.0,
                    "q2": 0.0,
                    "q3": 0.0,
                    "q4": 0.0,
                },
            ],
        },
    )


@pytest.fixture
def xlsx_with_tools_sheet(tmp_path: Path) -> Path:
    """An ``.xlsx`` with a ``tools`` sheet, a ``wobjs`` sheet, and a traj sheet."""
    wb = Workbook()
    # traj sheet
    ws_traj = wb.active
    ws_traj.title = "Sheet1"  # type: ignore[union-attr]
    ws_traj.append(["x", "y", "z", "q1", "q2", "q3", "q4", "tool_index", "wobj_index"])  # type: ignore[union-attr]
    ws_traj.append([1.0, 2.0, 3.0, 1.0, 0.0, 0.0, 0.0, 0, 0])  # type: ignore[union-attr]
    ws_traj.append([4.0, 5.0, 6.0, 1.0, 0.0, 0.0, 0.0, 1, 0])  # type: ignore[union-attr]
    # tools sheet
    ws_tools = wb.create_sheet("tools")
    ws_tools.append(["name"])
    ws_tools.append(["Tool_A"])
    ws_tools.append(["Tool_B"])
    # wobjs sheet
    ws_wobjs = wb.create_sheet("wobjs")
    ws_wobjs.append(["name"])
    ws_wobjs.append(["Wobj_A"])
    path = tmp_path / "with_tools_sheet.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_meta_sheet(tmp_path: Path) -> Path:
    """An ``.xlsx`` with a ``meta`` sheet (key/value) and a traj sheet."""
    wb = Workbook()
    ws_traj = wb.active
    ws_traj.title = "Sheet1"  # type: ignore[union-attr]
    ws_traj.append(["x", "y", "z"])  # type: ignore[union-attr]
    ws_traj.append([1.0, 2.0, 3.0])  # type: ignore[union-attr]
    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["key", "value"])
    ws_meta.append(["name", "Traj_Test"])
    path = tmp_path / "with_meta_sheet.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_with_full_meta(tmp_path: Path) -> Path:
    """An ``.xlsx`` with a full ``meta`` sheet including robot_model and extra fields."""
    wb = Workbook()
    ws_traj = wb.active
    ws_traj.title = "Sheet1"  # type: ignore[union-attr]
    ws_traj.append(["x", "y", "z"])  # type: ignore[union-attr]
    ws_traj.append([1.0, 2.0, 3.0])  # type: ignore[union-attr]
    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["key", "value"])
    ws_meta.append(["name", "Trajectoire_Soudure"])
    ws_meta.append(["robot_model", "IRB6700-205/2.80"])
    ws_meta.append(["author", "Jean Dupont"])
    path = tmp_path / "full_meta.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def xlsx_empty_rows(tmp_path: Path) -> Path:
    """An ``.xlsx`` with 2 valid rows and 1 fully empty row."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["x", "y", "z"])  # type: ignore[union-attr]
    ws.append([1.0, 2.0, 3.0])  # type: ignore[union-attr]
    ws.append([None, None, None])  # empty row
    ws.append([4.0, 5.0, 6.0])  # type: ignore[union-attr]
    path = tmp_path / "empty_rows.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Fixtures — synthetic CSV files
# ---------------------------------------------------------------------------


@pytest.fixture
def csv_simple(tmp_path: Path) -> Path:
    """Standard comma-separated CSV with 2 complete trajectory rows."""
    p = tmp_path / "simple.csv"
    p.write_text(
        "x,y,z,q1,q2,q3,q4,move_type,speed,zone,tool,wobj\n"
        "100.0,200.0,300.0,1.0,0.0,0.0,0.0,MoveL,v500,z10,Tool_formage,Wobj_SerreFlan\n"
        "150.0,250.0,350.0,1.0,0.0,0.0,0.0,MoveL,v500,z10,Tool_formage,Wobj_SerreFlan\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def csv_semicolon(tmp_path: Path) -> Path:
    """Semicolon-separated CSV (French Excel export style)."""
    p = tmp_path / "semicolon.csv"
    p.write_text(
        "x;y;z;q1;q2;q3;q4\n"
        "10.0;20.0;30.0;1.0;0.0;0.0;0.0\n"
        "40.0;50.0;60.0;1.0;0.0;0.0;0.0\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def csv_missing_xyz(tmp_path: Path) -> Path:
    """A CSV with only quaternion columns — XYZ missing."""
    p = tmp_path / "missing_xyz.csv"
    p.write_text("q1,q2,q3,q4\n1,0,0,0\n", encoding="utf-8")
    return p


@pytest.fixture
def csv_xyz_only(tmp_path: Path) -> Path:
    """A CSV with only XYZ columns — quaternions absent."""
    p = tmp_path / "xyz_only.csv"
    p.write_text("x,y,z\n1.0,2.0,3.0\n", encoding="utf-8")
    return p


@pytest.fixture
def csv_with_bom(tmp_path: Path) -> Path:
    """A UTF-8 BOM CSV file (as produced by Excel 'Save as CSV')."""
    p = tmp_path / "bom.csv"
    p.write_text("x,y,z\n1.0,2.0,3.0\n", encoding="utf-8-sig")
    return p


@pytest.fixture
def csv_aliases(tmp_path: Path) -> Path:
    """A CSV using column aliases (PosX, vitesse) instead of canonical names."""
    p = tmp_path / "aliases.csv"
    p.write_text(
        "PosX,PosY,PosZ,q1,q2,q3,q4,vitesse\n1.0,2.0,3.0,1.0,0.0,0.0,0.0,v500\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def csv_with_tools(tmp_path: Path) -> Path:
    """A CSV with inline ``tool`` and ``wobj`` columns."""
    p = tmp_path / "with_tools.csv"
    p.write_text(
        "x,y,z,q1,q2,q3,q4,tool,wobj\n"
        "1.0,2.0,3.0,1.0,0.0,0.0,0.0,Tool_A,Wobj_A\n"
        "4.0,5.0,6.0,1.0,0.0,0.0,0.0,Tool_B,Wobj_A\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def csv_empty_rows(tmp_path: Path) -> Path:
    """A CSV with 2 valid rows and 1 fully empty row."""
    p = tmp_path / "empty_rows.csv"
    p.write_text(
        "x,y,z\n1.0,2.0,3.0\n,,\n4.0,5.0,6.0\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def csv_full(tmp_path: Path) -> Path:
    """A CSV with all columns present including MoveL and MoveJ."""
    p = tmp_path / "full.csv"
    p.write_text(
        "x,y,z,q1,q2,q3,q4,move_type,speed,zone,tool,wobj\n"
        "1.0,2.0,3.0,1.0,0.0,0.0,0.0,MoveL,v500,z10,Tool_A,Wobj_A\n"
        "4.0,5.0,6.0,1.0,0.0,0.0,0.0,MoveJ,v250,z5,Tool_A,Wobj_A\n",
        encoding="utf-8",
    )
    return p


# ---------------------------------------------------------------------------
# Fixtures — Trajectory DataFrames and metadata (for test_trajectory.py)
# ---------------------------------------------------------------------------


@pytest.fixture
def minimal_meta() -> TrajectoryMeta:
    """Minimal ``TrajectoryMeta`` with only a name."""
    return TrajectoryMeta(name="test_traj")


@pytest.fixture
def minimal_df() -> pd.DataFrame:
    """Minimal ``DataFrame`` with all 7 required columns."""
    return pd.DataFrame(
        {
            "x": [100.0, 150.0],
            "y": [200.0, 250.0],
            "z": [300.0, 350.0],
            "q1": [1.0, 1.0],
            "q2": [0.0, 0.0],
            "q3": [0.0, 0.0],
            "q4": [0.0, 0.0],
        }
    )


@pytest.fixture
def complete_df() -> pd.DataFrame:
    """Complete ``DataFrame`` with all CONVERTER_COLUMNS present."""
    return pd.DataFrame(
        {
            "x": [100.0, 150.0],
            "y": [200.0, 250.0],
            "z": [300.0, 350.0],
            "q1": [1.0, 1.0],
            "q2": [0.0, 0.0],
            "q3": [0.0, 0.0],
            "q4": [0.0, 0.0],
            "cf1": [0, 0],
            "cf4": [0, 0],
            "cf6": [0, 0],
            "cfx": [0, 0],
            "move_type": ["MoveL", "MoveL"],
            "speed": ["v500", "v500"],
            "zone": ["z10", "z10"],
            "tool_index": [0, 0],
            "wobj_index": [0, 0],
        }
    )


@pytest.fixture
def complete_df_with_eax() -> pd.DataFrame:
    """Complete ``DataFrame`` with an active ``eax_a`` column."""
    return pd.DataFrame(
        {
            "x": [100.0],
            "y": [200.0],
            "z": [300.0],
            "q1": [1.0],
            "q2": [0.0],
            "q3": [0.0],
            "q4": [0.0],
            "cf1": [0],
            "cf4": [0],
            "cf6": [0],
            "cfx": [0],
            "move_type": ["MoveL"],
            "speed": ["v500"],
            "zone": ["z10"],
            "tool_index": [0],
            "wobj_index": [0],
            "eax_a": [45.0],
        }
    )


@pytest.fixture
def simple_trajectory(
    minimal_meta: TrajectoryMeta, complete_df: pd.DataFrame
) -> Trajectory:
    """A complete 2-point trajectory with one tool and one wobj."""
    return Trajectory(
        meta=minimal_meta,
        points=complete_df,
        tools=["Tool_formage"],
        wobjs=["Wobj_SerreFlan"],
    )


@pytest.fixture
def csv_unknown_col(tmp_path: Path) -> Path:
    """A CSV with one unknown column alongside valid XYZ columns."""
    p = tmp_path / "unknown_col.csv"
    p.write_text(
        "x,y,z,custom_col\n1.0,2.0,3.0,99\n",
        encoding="utf-8",
    )
    return p
